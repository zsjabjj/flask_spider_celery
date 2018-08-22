# -*- coding: utf-8 -*-
import datetime
import os

import openpyxl

from inchange_net.constants import MIDEA_DIR
from inchange_net.utils.DateRangeAndTimeStamp import dir_time


class SaveMidea(object):

    def __init__(self, date_time):
        '''初始化'''
        # 获取到指定日期
        self.date_time = date_time
        # 判断指定日期是否是手动输入
        if self.date_time:
            pass
        else:
            self.date_time = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%m月%d日')
        # 获取所需的年月日
        Year, Month, Day, year_t, month_t, day_t = dir_time()
        # 设置存储目录
        # 前一天
        mideaDir_y = '{midea_dir}midea_{year}/midea_{month}/'.format(midea_dir=MIDEA_DIR, year=Year, month=Year + Month)

        # 判断文件夹是否存在
        if os.path.exists(mideaDir_y):
            pass
        else:
            # os.mkdir(mideaDir_y)
            os.makedirs(mideaDir_y)

        self.filename = '{midea_path}竞品日数据跟踪_{month}.xlsx'.format(midea_path=mideaDir_y, month=(datetime.date.today() - datetime.timedelta(days=1)).strftime('%Y%m'))

        if os.path.exists(self.filename):
            self.wb = openpyxl.load_workbook(self.filename)

        else:
            self.wb = openpyxl.Workbook()

        try:
            self.ws_cg = self.wb['炒锅']
            self.ws_jg = self.wb['煎锅']
            self.ws_ng = self.wb['奶锅']
            self.ws_tg = self.wb['汤锅']
            self.ws_zg = self.wb['蒸锅']
        except:
            self.ws_cg = self.wb.active
            self.ws_jg = self.wb.create_sheet()
            self.ws_ng = self.wb.create_sheet()
            self.ws_tg = self.wb.create_sheet()
            self.ws_zg = self.wb.create_sheet()
            # self.ws = self.wb.active
            self.ws_cg.title = '炒锅'
            self.ws_jg.title = '煎锅'
            self.ws_ng.title = '奶锅'
            self.ws_tg.title = '汤锅'
            self.ws_zg.title = '蒸锅'

            self.ws_cg = self.wb['炒锅']
            self.ws_jg = self.wb['煎锅']
            self.ws_ng = self.wb['奶锅']
            self.ws_tg = self.wb['汤锅']
            self.ws_zg = self.wb['蒸锅']

            for ws in [self.ws_cg, self.ws_jg, self.ws_ng, self.ws_tg, self.ws_zg]:
                ws.cell(row=1, column=1, value='日期1')
                ws.cell(row=1, column=2, value='行业-访客数')
                ws.cell(row=1, column=3, value='行业-支付件数')
                ws.cell(row=1, column=4, value='行业-客单价')
                ws.cell(row=1, column=5, value='行业-搜索点击人数')
                ws.cell(row=1, column=6, value='日期2')
                ws.cell(row=1, column=7, value='美的-访客数')
                # ws.cell(row=1, column=8, value='美的-支付件数')
                ws.cell(row=1, column=8, value='美的-客单价')
                ws.cell(row=1, column=9, value='美的-支付转化率')
                ws.cell(row=1, column=10, value='美的-搜索点击人数')
                ws.cell(row=1, column=11, value='美的-支付件数')
                ws.cell(row=1, column=12, value='日期2')
                ws.cell(row=1, column=13, value='苏泊尔-访客数')

                ws.cell(row=1, column=14, value='苏泊尔-客单价')
                ws.cell(row=1, column=15, value='苏泊尔-支付转化率')
                ws.cell(row=1, column=16, value='苏泊尔-搜索点击人数')
                ws.cell(row=1, column=17, value='苏泊尔-支付件数')
                ws.cell(row=1, column=18, value='日期2')
                ws.cell(row=1, column=19, value='炊大皇-访客数')

                ws.cell(row=1, column=20, value='炊大皇-客单价')
                ws.cell(row=1, column=21, value='炊大皇-支付转化率')
                ws.cell(row=1, column=22, value='炊大皇-搜索点击人数')
                ws.cell(row=1, column=23, value='炊大皇-支付件数')
                ws.cell(row=1, column=24, value='日期2')
                ws.cell(row=1, column=25, value='爱仕达-访客数')

                ws.cell(row=1, column=26, value='爱仕达-客单价')
                ws.cell(row=1, column=27, value='爱仕达-支付转化率')
                ws.cell(row=1, column=28, value='爱仕达-搜索点击人数')
                ws.cell(row=1, column=29, value='爱仕达-支付件数')
                if self.ws_cg == ws:
                    ws.cell(row=1, column=30, value='日期2')
                    ws.cell(row=1, column=31, value='臻三环-访客数')

                    ws.cell(row=1, column=32, value='臻三环-客单价')
                    ws.cell(row=1, column=33, value='臻三环-支付转化率')
                    ws.cell(row=1, column=34, value='臻三环-搜索点击人数')
                    ws.cell(row=1, column=35, value='臻三环-支付件数')
                    ws.cell(row=1, column=36, value='日期2')
                    ws.cell(row=1, column=37, value='feillers-访客数')

                    ws.cell(row=1, column=28, value='feillers-客单价')
                    ws.cell(row=1, column=39, value='feillers-支付转化率')
                    ws.cell(row=1, column=40, value='feillers-搜索点击人数')
                    ws.cell(row=1, column=41, value='feillers-支付件数')

        self.num_cg = 0
        self.num_jg = 0
        self.num_ng = 0
        self.num_tg = 0
        self.num_zg = 0

    def save_midea_data(self, titles):
        '''数据存储分类'''
        for _, item in enumerate(titles):
            # 判断数据类型
            if isinstance(item, dict):
                if '炒锅' == item['category']:

                    self.save_item(self.ws_cg, item, self.num_cg)
                    self.num_cg += 1
                elif '煎锅' == item['category']:

                    self.save_item(self.ws_jg, item, self.num_jg)
                    self.num_jg += 1
                elif '奶锅' == item['category']:

                    self.save_item(self.ws_ng, item, self.num_ng)
                    self.num_ng += 1
                elif '汤锅' == item['category']:

                    self.save_item(self.ws_tg, item, self.num_tg)
                    self.num_tg += 1
                elif '蒸锅' == item['category']:

                    self.save_item(self.ws_zg, item, self.num_zg)
                    self.num_zg += 1

    def save_item(self, ws, data, num):
        '''数据存储'''
        if data:

            if num:
                maxRow = ws.max_row

            else:
                maxRow = ws.max_row + 1

            final_date = [str(int(no)) for no in data['time_date'].split('-')]
            data_day = '/'.join(final_date)

            if 'dapan' == data['mark']:

                ws.cell(row=maxRow, column=1, value=data_day)
                ws.cell(row=maxRow, column=2, value=data['uv'])
                ws.cell(row=maxRow, column=3, value=data['payItemQty'])
                ws.cell(row=maxRow, column=4, value=data['payPct'])
                ws.cell(row=maxRow, column=5, value=data['searchUvCnt'])

            else:

                ws.cell(row=maxRow, column=6, value=data_day)
                ws.cell(row=maxRow, column=7, value=data[30652]['uv'])

                ws.cell(row=maxRow, column=8, value=data[30652]['payPct'])
                ws.cell(row=maxRow, column=9, value=data[30652]['payRate'])
                ws.cell(row=maxRow, column=10, value=data[30652]['searchUvCnt'])
                ws.cell(row=maxRow, column=11, value=data[30652]['payItemQty'])
                ws.cell(row=maxRow, column=12, value=data_day)
                ws.cell(row=maxRow, column=13, value=data[30844]['uv'])

                ws.cell(row=maxRow, column=14, value=data[30844]['payPct'])
                ws.cell(row=maxRow, column=15, value=data[30844]['payRate'])
                ws.cell(row=maxRow, column=16, value=data[30844]['searchUvCnt'])
                ws.cell(row=maxRow, column=17, value=data[30844]['payItemQty'])
                ws.cell(row=maxRow, column=18, value=data_day)
                ws.cell(row=maxRow, column=19, value=data[5725958]['uv'])

                ws.cell(row=maxRow, column=20, value=data[5725958]['payPct'])
                ws.cell(row=maxRow, column=21, value=data[5725958]['payRate'])
                ws.cell(row=maxRow, column=22, value=data[5725958]['searchUvCnt'])
                ws.cell(row=maxRow, column=23, value=data[5725958]['payItemQty'])
                ws.cell(row=maxRow, column=24, value=data_day)
                ws.cell(row=maxRow, column=25, value=data[3222885]['uv'])

                ws.cell(row=maxRow, column=26, value=data[3222885]['payPct'])
                ws.cell(row=maxRow, column=27, value=data[3222885]['payRate'])
                ws.cell(row=maxRow, column=28, value=data[3222885]['searchUvCnt'])
                ws.cell(row=maxRow, column=29, value=data[3222885]['payItemQty'])
                if '炒锅' == data['category']:
                    ws.cell(row=maxRow, column=30, value=data_day)
                    ws.cell(row=maxRow, column=31, value=data[531544226]['uv'])

                    ws.cell(row=maxRow, column=32, value=data[531544226]['payPct'])
                    ws.cell(row=maxRow, column=33, value=data[531544226]['payRate'])
                    ws.cell(row=maxRow, column=34, value=data[531544226]['searchUvCnt'])
                    ws.cell(row=maxRow, column=35, value=data[531544226]['payItemQty'])
                    ws.cell(row=maxRow, column=36, value=data_day)
                    ws.cell(row=maxRow, column=37, value=data[1344359932]['uv'])

                    ws.cell(row=maxRow, column=38, value=data[1344359932]['payPct'])
                    ws.cell(row=maxRow, column=39, value=data[1344359932]['payRate'])
                    ws.cell(row=maxRow, column=40, value=data[1344359932]['searchUvCnt'])
                    ws.cell(row=maxRow, column=41, value=data[1344359932]['payItemQty'])

            self.wb.save(self.filename)
