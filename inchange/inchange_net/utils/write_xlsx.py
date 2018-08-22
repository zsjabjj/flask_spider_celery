# -*- coding: utf-8 -*-
import openpyxl
import logging
from openpyxl.styles import Alignment, Border, Side
import datetime
import csv
import codecs
from openpyxl.utils import get_column_letter


def save_another(filename, upload_dir):
    # 标记必填有空单元格
    mark_num = 0
    # # 标记必填单元格数量，用来判断达标18个
    # markNum = 0
    shop_name = filename.split('.')[0].split('/')[-1]
    
    logging.info(filename)
    item_dict = dict()
    with open(filename, 'r', encoding='gbk') as csvfile:
        # 读取csv文件内容为生成器
        readers = csv.DictReader(csvfile)
        num = 0
        nick_sum_bcd = 0
        nick_sum = 0
        dis_sum_bcd = 0
        dis_sum = 0
        

        # 遍历生成器为列表
        row_list = [row for row in readers]
        # print('row_list:', type(row_list))
        # print('产品名称' in row_list)

        # 遍历列表，获得元素下标以及元素
        for ind, row in enumerate(row_list):

            # print(ind, '---->', type(list(row)))
            # print('=======', list(row))
            # print('宝贝标题 ' in list(row))

            # 区分荣事达官旗和分销
            if '宝贝标题 ' in list(row):
                item_dict['shop_name'] = '荣事达官旗'
                # 荣事达官旗
                # if row['订单状态'] not in '退款成功,已退款,交易关闭,等待付款':
                if '.' in row['买家实际支付金额']:
                    price_int = int(''.join(row['买家实际支付金额'].split('.')))
                    price = price_int / (10 ** len(row['买家实际支付金额'].split('.')[1]))
                else:
                    price = int(row['买家实际支付金额'])
                # print(type(row['买家实际支付金额']))
                if 'BCD' in row['宝贝标题 ']:

                    # 冰箱
                    # nick_sum_bcd += row['买家实际支付金额']

                    nick_sum_bcd += price
                    pass
                else:
                    # 洗衣机
                    nick_sum += price

                    pass

                pass
            else:
                # 统一处理
                # print('yuan:', row)

                # 判断采购状态以及分销商会员名是否存在
                # if row['采购单状态'] not in '退款成功,已退款,交易关闭,等待付款' and '-' == row['分销商会员名']:
                if '-' == row['分销商会员名']:

                    if '.' in row['分销商销售收入']:
                        price_int = int(''.join(row['分销商销售收入'].split('.')))
                        price = price_int / (10 ** len(row['分销商销售收入'].split('.')[1]))
                    else:
                        price = int(row['分销商销售收入'])

                    # 计数，可以一个用户买两件商品
                    num += 1
                    # 遍历key，将主信息填入分信息中
                    for key in row.keys():
                        # print('key', row[key])
                        # if row_list[ind - num][key] is not '-':
                        if row[key] is '-':
                            row[key] = row_list[ind - num][key]
                    # print('if:', row)

                    

                    # if '三洋' in row['分销商会员名'] or '帝度' in row['分销商会员名'] or '惠而浦' in row['分销商会员名'] or '荣事达' == shop_name:
                        # 统计官旗和分销店的销售收入
                    if '旗舰店' in row['分销商会员名'] and ('三洋' in row['分销商会员名'] or '帝度' in row['分销商会员名'] or '惠而浦' in row['分销商会员名']):
                        # if '三洋' in row['分销商会员名'] or '帝度' in row['分销商会员名'] or '惠而浦' in row['分销商会员名']:
                        if '三洋' in row['产品名称']:
                            item_dict['shop_name'] = '三洋'
                        elif '帝度' in row['产品名称']:
                            item_dict['shop_name'] = '帝度'
                        elif '惠而浦' in row['产品名称']:
                            item_dict['shop_name'] = '惠而浦'
                        elif '荣事达' in row['产品名称']:
                            item_dict['shop_name'] = '荣事达官旗'
                        # print(row['分销商会员名'])
                        if 'BCD' in row['产品名称']:
                            # 冰箱
                            # print('qijianbcd:', row['分销商销售收入'], row['收件人'], ind)
                            # nick_sum_bcd += int(''.join(row['分销商销售收入'].split('.')))
                            nick_sum_bcd += price
                            pass
                        else:
                            # 洗衣机
                            # print('qijian:', row['分销商销售收入'], row['收件人'], ind)
                            nick_sum += price
                            pass
                        # nick_sum += int(''.join(row['分销商销售收入'].split('.')))
                    else:
                        if '三洋' in row['产品名称']:
                            item_dict['shop_name'] = '三洋'
                        elif '帝度' in row['产品名称']:
                            item_dict['shop_name'] = '帝度'
                        elif '惠而浦' in row['产品名称']:
                            item_dict['shop_name'] = '惠而浦'
                        elif '荣事达' in row['产品名称']:
                            item_dict['shop_name'] = '荣事达分销'

                        if 'BCD' in row['产品名称']:
                            # print('elif', row['分销商会员名'])
                            # 分销冰箱
                            # print('fenxiaobcd:', row['分销商销售收入'], row['收件人'], ind)
                            dis_sum_bcd += price
                            pass
                        else:
                            # print('else', row['分销商会员名'])
                            # 分销洗衣机
                            # print('fenxiao:', row['分销商销售收入'], row['收件人'], ind)
                            dis_sum += price
                            pass
                        # dis_sum += int(''.join(row['分销商销售收入'].split('.')))
                else:
                    num = 0

    # print('nick sum:%s, dis sum:%s, sum:%s' % (round(nick_sum / 1000000, 1), round(dis_sum / 1000000, 1), round(nick_sum / 1000000 + dis_sum / 1000000, 1)))

    # print(item_dict['shop_name'])
    # print('nick_sum:', nick_sum)

    

    # print('sSumBCD:', sSumBCD)
    # print('sSum:', sSum)
    results_dict = dict()
    try:
        results_dict['品牌'] = item_dict['shop_name']
        
    except:
        if '三洋' == shop_name:
            results_dict['品牌'] = '三洋'           
            pass
        elif '帝度' == shop_name:
            results_dict['品牌'] = '帝度'
            pass
        elif '惠而浦' == shop_name:
            results_dict['品牌'] = '惠而浦'
            pass
        elif '荣事达' == shop_name:
            results_dict['品牌'] = '荣事达分销'
            pass
        elif '荣事达官旗' == shop_name:
            results_dict['品牌'] = '荣事达官旗'
            pass
        nick_sum_bcd = 0
        nick_sum = 0
        dis_sum_bcd = 0
        dis_sum = 0

    # if nick_sum_bcd:
    results_dict['BCD官旗销售额'] = nick_sum_bcd
    # else:
        # results_dict['BCD官旗销售额'] = 0.000000001
    # if nick_sum:
    results_dict['官旗销售额'] = nick_sum
    # else:
        # results_dict['官旗销售额'] = 0.000000001
    # if dis_sum_bcd:
    results_dict['BCD分销商销售额'] = dis_sum_bcd
    # else:
        # results_dict['BCD分销商销售额'] = 0.000000001
    # if dis_sum:
    results_dict['分销商销售额'] = dis_sum
    # else:
        # results_dict['分销商销售额'] = 0.000000001

    sSumBCD = nick_sum_bcd + dis_sum_bcd
    sSum = nick_sum + dis_sum
    # results_dict['官旗销售额'] = nick_sum
    # results_dict['BCD分销商销售额'] = dis_sum_bcd
    # results_dict['分销商销售额'] = dis_sum
    results_dict['BCD两者总和'] = sSumBCD
    results_dict['两者总和'] = sSum
    # print(results_dict)

    recent_time = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%m月%d日')

    '''
    5月销售任务4000万，5月20日销售达成150.2万，官旗完成104.8万，占总体销售69.80%，分销完成45.4万，占总体销售30.20%；
    5月累计销售达成1416.2万，累计进度35.41%，本月31天，时间进度64.52%，其中三洋当日完成126.7万，累计完成1207.3万，惠而浦17.7万，累计完成114.5万，帝度0.3万，累计完成7.7万，荣事达5.5万，累计完成86.6万，以上请查阅。
    '''

    # 存文件
    # wb = openpyxl.load_workbook('01-5月日报-三洋&惠而浦&帝度&荣事达销售数据.xlsx')
    # wb = openpyxl.load_workbook('11_20180523_143722.xlsx')
    try:
        wb = openpyxl.load_workbook(upload_dir + '01-5月日报-三洋&惠而浦&帝度&荣事达销售数据_%s.xlsx' % datetime.datetime.now().strftime('%Y%m%d'))
    except:
        for i in range(1, 11):
            file_name = upload_dir + '01-5月日报-三洋&惠而浦&帝度&荣事达销售数据_%s.xlsx' % (datetime.date.today() - datetime.timedelta(days=i)).strftime('%Y%m%d')
            try:
                wb = openpyxl.load_workbook(file_name)
            except Exception as err:
                logging.error(err)
                continue
            else:
                break
        # wb = openpyxl.load_workbook(upload_dir + '01-5月日报-三洋&惠而浦&帝度&荣事达销售数据_%s.xlsx' % (datetime.date.today() - datetime.timedelta(days=1)).strftime('%Y%m%d'))
    # sheet = wb.get_sheet_by_name()
    sheet = wb.active
    # print(wb.properties)

    sheet_title = sheet.title
    title_time = sheet_title.split('_')[1]
    # print(sheet_title)
    # print(title_time == datetime.datetime.now().strftime('%Y%m%d'))


    a1 = sheet['A1'].value
    a2 = sheet['A2'].value
    # print(a1)
    # print(a2)
    # print(sheet.max_row)
    # print(sheet.max_column)

    maxRow = sheet.max_row
    maxCol = sheet.max_column

    # 判断必填单元格B(4-12, 14-22)是否已经填完
    for i in range(4, 23):
        if i == 13:
            pass
        elif sheet.cell(row=i, column=maxCol).value is None:
            mark_num += 1
            pass
        # else:
            # mark_num += 1
    # 如果mark_num有值，说明已经填写过，但是没有填写完全
    logging.info(mark_num)
    # if mark_num and title_time == datetime.datetime.now().strftime('%Y%m%d'):
    if title_time == datetime.datetime.now().strftime('%Y%m%d'):
        # print('haha')
        maxCol -= 1
    # for i in range(1, maxCol + 1):
    #     if not sheet.cell(row=4, column=i).value:
    #         print(i)
    else:
        left, right, top, bottom = [Side(style='thin', color='000000')] * 4
        sheet.unmerge_cells(start_row=1, start_column=1, end_row=1, end_column=maxCol)
        sheet.unmerge_cells(start_row=2, start_column=1, end_row=2, end_column=maxCol)
        # 标题
        sc_title = sheet.cell(row=1, column=1, value=a1)
        sc_title.alignment = Alignment(horizontal='center', vertical='center')
        sc_title.border = Border(left=left, right=right, top=top, bottom=bottom)
        # sheet.merge_cells('B1:G1') # 合并一行中的几个单元格
        # 合并单元格

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=maxCol + 1)
        # 活动
        sc_act = sheet.cell(row=2, column=1, value=a2)
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=maxCol + 1)
        sc_act.alignment = Alignment(horizontal='center', vertical='center')
        sc_act.border = Border(left=left, right=right, top=top, bottom=bottom)

        sc1 = sheet.cell(row=3, column=maxCol + 1, value=recent_time)
        sc1.alignment = Alignment(horizontal='center', vertical='center')
        sc1.border = Border(left=left, right=right, top=top, bottom=bottom)


        # 处理文本正文
        for i in range(4, maxRow - 9):
            # print(i)
            sc = sheet.cell(row=i, column=maxCol + 1)
            sc.alignment = Alignment(horizontal='center', vertical='center')
            sc.border = Border(left=left, right=right, top=top, bottom=bottom)

        sheet.title = '总表_%s' % datetime.datetime.now().strftime('%Y%m%d')
        sheet_title = sheet.title
        # print(sheet_title)

    if '三洋' == results_dict['品牌']:
        # sheet.cell(row=4, column=maxCol + 1, value=round(results_dict['官旗销售额'] / 10000, 1))
        sheet.cell(row=4, column=maxCol + 1, value=results_dict['官旗销售额'])
        # sheet.cell(row=5, column=maxCol + 1, value=round(results_dict['分销商销售额'] / 10000, 1))
        # sheet.cell(row=6, column=maxCol + 1, value=round(results_dict['两者总和'] / 10000, 1))
        sheet.cell(row=5, column=maxCol + 1, value=results_dict['分销商销售额'])
        sheet.cell(row=6, column=maxCol + 1, value=results_dict['两者总和'])

    elif '帝度' == results_dict['品牌']:
        sheet.cell(row=17, column=maxCol + 1, value=results_dict['BCD官旗销售额'])
        # sheet.cell(row=17, column=maxCol + 1, value=round(results_dict['BCD官旗销售额'] / 10000, 1))
        # sheet.cell(row=18, column=maxCol + 1, value=round(results_dict['BCD分销商销售额'] / 10000, 1))
        # sheet.cell(row=19, column=maxCol + 1, value=round(results_dict['BCD两者总和'] / 10000, 1))
        sheet.cell(row=18, column=maxCol + 1, value=results_dict['BCD分销商销售额'])
        sheet.cell(row=19, column=maxCol + 1, value=results_dict['BCD两者总和'])

    elif '惠而浦' == results_dict['品牌']:
        # sheet.cell(row=7, column=maxCol + 1, value=round(results_dict['官旗销售额'] / 10000, 1))
        sheet.cell(row=7, column=maxCol + 1, value=results_dict['官旗销售额'])
        sheet.cell(row=8, column=maxCol + 1, value=results_dict['分销商销售额'])
        sheet.cell(row=9, column=maxCol + 1, value=results_dict['两者总和'])
        sheet.cell(row=14, column=maxCol + 1, value=results_dict['BCD官旗销售额'])
        sheet.cell(row=15, column=maxCol + 1, value=results_dict['BCD分销商销售额'])
        sheet.cell(row=16, column=maxCol + 1, value=results_dict['BCD两者总和'])
        # sheet.cell(row=8, column=maxCol + 1, value=round(results_dict['分销商销售额'] / 10000, 1))
        # sheet.cell(row=9, column=maxCol + 1, value=round(results_dict['两者总和'] / 10000, 1))
        # sheet.cell(row=14, column=maxCol + 1, value=round(results_dict['BCD官旗销售额'] / 10000, 1))
        # sheet.cell(row=15, column=maxCol + 1, value=round(results_dict['BCD分销商销售额'] / 10000, 1))
        # sheet.cell(row=16, column=maxCol + 1, value=round(results_dict['BCD两者总和'] / 10000, 1))

    elif '荣事达官旗' == results_dict['品牌']:
        sheet.cell(row=10, column=maxCol + 1, value=results_dict['官旗销售额'])
        # sheet.cell(row=10, column=maxCol + 1, value=round(results_dict['官旗销售额'] / 10000, 1))
        # sheet.cell(row=20, column=maxCol + 1, value=round(results_dict['BCD官旗销售额'] / 10000, 1))
        sheet.cell(row=20, column=maxCol + 1, value=results_dict['BCD官旗销售额'])
        # sheet.cell(row=18, column=maxCol + 1, value=results_dict['BCD分销商销售额'])
        # sheet.cell(row=19, column=maxCol + 1, value=results_dict['BCD两者总和'])

    elif '荣事达分销' == results_dict['品牌']:
        sheet.cell(row=11, column=maxCol + 1, value=results_dict['分销商销售额'])
        sheet.cell(row=21, column=maxCol + 1, value=results_dict['BCD分销商销售额'])
        # sheet.cell(row=11, column=maxCol + 1, value=round(results_dict['分销商销售额'] / 10000, 1))
        # sheet.cell(row=21, column=maxCol + 1, value=round(results_dict['BCD分销商销售额'] / 10000, 1))
    try:
        rsd = sheet.cell(row=10, column=maxCol + 1).value * sheet.cell(row=11, column=maxCol + 1).value * sheet.cell(row=20, column=maxCol + 1).value * sheet.cell(row=21, column=maxCol + 1).value
    except Exception as err:
        logging.error(err)

    # if rsd:
    else:
        rsd_sum = sheet.cell(row=10, column=maxCol + 1).value + sheet.cell(row=11, column=maxCol + 1).value
        rsd_bcd = sheet.cell(row=20, column=maxCol + 1).value + sheet.cell(row=21, column=maxCol + 1).value
        sheet.cell(row=12, column=maxCol + 1, value=rsd_sum)
        sheet.cell(row=22, column=maxCol + 1, value=rsd_bcd)

    logging.info(mark_num)
    # if mark_num == 4 or mark_num == 3 and title_time == datetime.datetime.now().strftime('%Y%m%d'):
    if mark_num == 3 and title_time == datetime.datetime.now().strftime('%Y%m%d'):

        # 洗衣机总计
        WSum = sheet.cell(row=6, column=maxCol + 1).value + sheet.cell(row=9, column=maxCol + 1).value + sheet.cell(row=12, column=maxCol + 1).value
        sheet.cell(row=13, column=maxCol + 1, value=WSum)
        # 冰箱总计
        BCDSum = sheet.cell(row=16, column=maxCol + 1).value + sheet.cell(row=19, column=maxCol + 1).value + sheet.cell(row=22, column=maxCol + 1).value
        sheet.cell(row=23, column=maxCol + 1, value=BCDSum)
        # 三洋总计
        sySum = sheet.cell(row=6, column=maxCol + 1).value
        # sheet.cell(row=24, column=maxCol + 1, value=round(sySum / 10000, 1))
        sheet.cell(row=24, column=maxCol + 1, value=sySum)
        # 惠而浦总计
        hrpSum = sheet.cell(row=9, column=maxCol + 1).value + sheet.cell(row=16, column=maxCol + 1).value
        # sheet.cell(row=25, column=maxCol + 1, value=round(hrpSum / 10000, 1))
        sheet.cell(row=25, column=maxCol + 1, value=hrpSum)
        # 帝度总计
        ddSum = sheet.cell(row=19, column=maxCol + 1).value
        # sheet.cell(row=26, column=maxCol + 1, value=round(ddSum / 10000, 1))
        sheet.cell(row=26, column=maxCol + 1, value=ddSum)
        # 荣事达总计
        rsdSum = sheet.cell(row=12, column=maxCol + 1).value + sheet.cell(row=22, column=maxCol + 1).value
        # sheet.cell(row=27, column=maxCol + 1, value=round(rsdSum / 10000, 1))
        sheet.cell(row=27, column=maxCol + 1, value=rsdSum)
        # 总计
        total = sySum + hrpSum + ddSum + rsdSum
        # sheet.cell(row=28, column=maxCol + 1, value=round(total / 10000, 1))
        sheet.cell(row=28, column=maxCol + 1, value=total)
        # 官旗合计
        gqSum = 0
        for gq in range(4, 21, 3):
            if gq > 10:
                gq += 1
            gqSum += sheet.cell(row=gq, column=maxCol + 1).value
        # sheet.cell(row=29, column=maxCol + 1, value=round(gqSum / 10000, 1))
        sheet.cell(row=29, column=maxCol + 1, value=gqSum)
        # 分销合计
        fxSum = 0
        for fx in range(5, 22, 3):
            if fx > 11:
                fx += 1
            fxSum += sheet.cell(row=fx, column=maxCol + 1).value
        logging.info(fxSum)
        # sheet.cell(row=31, column=maxCol + 1, value=round(fxSum / 10000, 1))
        sheet.cell(row=31, column=maxCol + 1, value=fxSum)
        # 官旗占比
        gq_ratio = round(gqSum / total * 100, 2)
        # gq_ratio = gqSum / total
        gqRatio = str(gq_ratio) + '%'
        sheet.cell(row=30, column=maxCol + 1, value=gqRatio)
        # sheet.cell(row=30, column=maxCol + 1, value=gq_ratio)
        # 分销占比
        fx_ratio = round(fxSum / total * 100, 2)
        # fx_ratio = fxSum / total
        fxRatio = str(fx_ratio) + '%'
        sheet.cell(row=32, column=maxCol + 1, value=fxRatio)
        # sheet.cell(row=32, column=maxCol + 1, value=fx_ratio)

        # 累计完成


        # 总结
        # sheet.unmerge_cells(start_row=36, start_column=1, end_row=42, end_column=4)

        # data = '5月销售任务4000万，{recent_time}销售达成{total}万，官旗完成{gqSum}万，占总体销售{gqRatio}，分销完成{fxSum}万，占总体销售{fxRatio}；5月累计销售达成1416.2万，累计进度35.41%，本月31天，时间进度64.52%，其中三洋当日完成126.7万，累计完成1207.3万，惠而浦17.7万，累计完成114.5万，帝度0.3万，累计完成7.7万，荣事达5.5万，累计完成86.6万，以上请查阅。'.format(recent_time=recent_time, total=total, gqSum=round(gqSum / 10000, 1), gqRatio=gqRatio, fxSum=round(fxSum / 10000, 1), fxRatio=fxRatio, )

        # sheet.cell(row=36, column=1, value=data)
        # sheet.merge_cells(start_row=36, start_column=1, end_row=42, end_column=4)

        # sheet.cell(row=36, column=1, value='')




        pass
    for i in range(7, maxCol - 1):
        sheet.column_dimensions[get_column_letter(i)].width = 0
    wb.save(upload_dir + '01-5月日报-三洋&惠而浦&帝度&荣事达销售数据_%s.xlsx' % datetime.datetime.now().strftime('%Y%m%d'))
    # # wb.save('11_20180523_143722.xlsx')
if __name__ == '__main__':
    save_another('采购单列表_201805221041.csv', '')
    save_another('采购单列表_201805221443.csv', '')
    save_another('采购单列表_201805221447.csv', '')
    save_another('采购单列表_201805231155.csv', '')
    save_another('ExportOrderList201805221453.csv', '')
    # # 向已存在的文件中写入数据
    # # 打开文件
    # file_path = "01-5月日报-三洋&惠而浦&帝度&荣事达销售数据.xlsx"
    # workbook = open_workbook(file_path)
    # # 获取sheet内容:根据sheet索引或者名称
    # sheet = workbook.sheet_by_index(0)  # 第一个sheet
    # # 复制上面打开的excel文件
    # new_workbook = copy(workbook)
    # # 从复制的excel文件中，得到写入用列的sheet表，这里写入第一个表即为0
    # new_sheet = new_workbook.get_sheet(0)
    # # 写入内容
    # new_sheet.write(0, 1, '10')
    # # 保存文件
    # book.save('simple2.xls')
    # new_workbook.save(file_path)
    # book.save(TemporaryFile())


    # result_filename = filename.split('.')[0]+'_result.csv'
    # with codecs.open(result_filename, 'w', 'utf_8_sig') as csvfile:
    #
    #     fieldnames = ['名称', '销售额']
    #     writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
    #     writer.writeheader()
    #     for key, value in results_dict.items():
    #         writer.writerow({'名称': key, '销售额': value})
    # return result_filename.split('/')[-1]
