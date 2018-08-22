# -*- coding: utf-8 -*-
import calendar
import xlrd
import xlwt
import openpyxl
import datetime
import re


def xls_extract(filename):
    '''处理商品效果文件'''
    # xls的日期处理
    yesterday = re.findall(r'.*商品效果-(.*)', filename.split('.')[0])[0]
    ytd_list = yesterday.split('-')
    year = int(ytd_list[0])
    month = int(ytd_list[1])
    day = int(ytd_list[2])
    _, endDay = calendar.monthrange(year, month)
    if day == endDay:
        next_day = datetime.date(year=year, month=month + 1, day=1).strftime('%Y%m%d')
    else:
        next_day = datetime.date(year=year, month=month, day=day + 1).strftime('%Y%m%d')

    titleList = ['日期',	'型号', '浏览量',	'访客数', '下单转化率', '支付转化率', '下单商品件数', '支付金额', '支付商品件数',	'加购件数', '收藏人数', '支付买家数', '客单价']
    titleNum = len(titleList)
    today = datetime.date(year=year, month=month, day=day).strftime('%m月%d日')

    wb = xlrd.open_workbook(filename)
    wb_new = xlwt.Workbook()

    sheets = wb.sheet_names()
    ws = wb.sheet_by_name(sheets[0])
    ws_add = wb_new.add_sheet('商品效果提取后的数据')
    # 设置单元格中内容位置 居中 居左右等
    # 创建格式对象
    alignment = xlwt.Alignment()  # Create Alignment
    # 水平居中
    alignment.horz = xlwt.Alignment.HORZ_CENTER
    # 垂直居中
    alignment.vert = xlwt.Alignment.VERT_CENTER
    # 创建style样式
    style = xlwt.XFStyle()  # Create Style
    # 将格式添加到样式中
    style.alignment = alignment  # Add Alignment to Style
    # 加样式 sheet.write(5, 5, 'Cell Contents', style)
    row_add = 0
    for colN in range(titleNum):

        ws_add.write(row_add, colN, titleList[colN], style)
    row_num = ws.nrows
    col_num = ws.ncols

    title_list = [a.value for a in ws.row(3)]
    index = title_list.index('商品在线状态')
    for i in range(4, row_num):
        if '当前在线' == ws.cell_value(i, index):

            row_add += 1
            # 日期
            ws_add.write(row_add, 0, today, style)
            # 型号
            # print(ws.cell(i, title_list.index('商品标题')).value.split(' '))
            model_list = ws.cell(i, title_list.index('商品标题')).value.split(' ')
            if 'Sanyo/三洋' == model_list[0]:
                # print(model_list[1])
                model = model_list[1]
            else:
                modelStr = re.findall(r'Sanyo/三洋(.*)', model_list[0])
                try:
                    # print(modelStr[0])
                    model = modelStr[0]
                except:
                    # print(model_list[0])
                    model = model_list[0]
            # ws_add.write(row_add, 1, ws.cell_value(i, title_list.index('商品标题')))
            ws_add.write(row_add, 1, model, style)
            # 各指标
            for j in range(2, titleNum):
                # ws_add.cell(row=row_add, column=j, value=ws.cell_value(i, title_list.index(titleList[j])))
                ws_add.write(row_add, j, ws.cell_value(i, title_list.index(titleList[j])), style)


    wb_new.save('%s_%s.xls' % (filename.split('.')[0], next_day))

if __name__ == '__main__':
    xls_extract('商品效果-2018-06-07.xls')
