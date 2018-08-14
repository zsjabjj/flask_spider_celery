# -*- coding: utf-8 -*-
import calendar
import datetime
import logging
import os

import openpyxl

from inchange_net.constants import SANYO_DIR


class SaveRatio(object):


    def __init__(self, date_time):
        '''初始化'''
        # 获取到指定日期
        self.date_time = date_time
        # 判断指定日期是否是手动输入
        if self.date_time:
            pass
        else:
            self.date_time = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%m月%d日')
        # 获取当天的一些日期信息
        year_t, month_t, day_t = datetime.datetime.today().strftime('%Y,%m,%d').split(',')
        # 获取前一天的一些日期信息
        year_y, month_y, day_y = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%Y,%m,%d').split(',')
        # 获取年
        Year = year_t if year_t == year_y else year_y
        # 获取月
        Month = month_t if month_t == month_y else month_y
        # 获取日
        Day = day_y
        # # 某年某月天数，返回是星期（0-6）和天数
        # _, days = calendar.monthrange(Year, Month)
        # 设置存储目录
        # 前一天
        sanyoDir_y = '{sanyo_dir}sanyo_{year}/sanyo_{month}/sanyo_{day}/'.format(sanyo_dir=SANYO_DIR, year=Year, month=Year + Month, day=Year + Month + Day)
        # 今天
        sanyoDir_t = '{sanyo_dir}sanyo_{year}/sanyo_{month}/sanyo_{day}/'.format(sanyo_dir=SANYO_DIR, year=year_t, month=year_t + month_t, day=year_t + month_t + day_t)
        # 判断文件夹是否存在
        if os.path.exists(sanyoDir_y):
            pass
        else:
            os.mkdir(sanyoDir_y)

        if os.path.exists(sanyoDir_t):
            pass
        else:
            os.mkdir(sanyoDir_t)

        # for i in range(1, 11):
        # day_time = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%Y%m%d')
        # 打开的文件名
        self.filename1 = '{sanyo_path}洗衣机行业占比累计数据_{day}.xlsx'.format(sanyo_path=sanyoDir_y, day=Year + Month + Day)
        self.filename2 = '{sanyo_path}冰箱行业占比累计数据_{day}.xlsx'.format(sanyo_path=sanyoDir_y, day=Year + Month + Day)
        # 保存的文件名
        self.filename3 = '{sanyo_path}洗衣机行业占比累计数据_{day}.xlsx'.format(sanyo_path=sanyoDir_t, day=year_t + month_t + day_t)
        self.filename4 = '{sanyo_path}冰箱行业占比累计数据_{day}.xlsx'.format(sanyo_path=sanyoDir_t, day=year_t + month_t + day_t)
        # self.filename1 = '/Users/tiger007/Desktop/666/ratio/sanyo/02-5月洗衣机行业占比累计数据_%s.xlsx' % (datetime.date.today() - datetime.timedelta(days=i)).strftime('%Y%m%d')
        # self.filename2 = '/Users/tiger007/Desktop/666/ratio/sanyo/03-5月冰箱行业占比累计数据_%s.xlsx' % (datetime.date.today() - datetime.timedelta(days=i)).strftime('%Y%m%d')

        try:

            self.wb_washer = openpyxl.load_workbook(self.filename1)
            self.wb_fridge = openpyxl.load_workbook(self.filename2)
        except Exception as err:
            logging.error(err)
            # continue
        else:
            sheet_name_washer = self.wb_washer.sheetnames[1]
            self.ws_washer = self.wb_washer[sheet_name_washer]
            self.ws_washer_add = self.wb_washer.create_sheet('全部原始数据')
            self.maxRow_washer = self.ws_washer.max_row
            self.row_washer = 1

            sheet_name_fridge = self.wb_fridge.sheetnames[1]
            self.ws_fridge = self.wb_fridge[sheet_name_fridge]
            self.ws_fridge_add = self.wb_fridge.create_sheet('全部原始数据')
            self.maxRow_fridge = self.ws_fridge.max_row
            self.row_fridge = 1

            # break


    # def __del__(self):
    #     self.wb_washer.save('./sanyo/02-5月洗衣机行业占比累计数据_%s.xlsx' % datetime.datetime.now().strftime('%Y%m%d'))
    #     self.wb_fridge.save('./sanyo/03-5月冰箱行业占比累计数据_%s.xlsx' % datetime.datetime.now().strftime('%Y%m%d'))
    #
    #     os.remove(self.filename1)
    #     os.remove(self.filename2)

    def save_ratio_data(self, titles):
        '''数据存储'''
        for title in titles:
            if type(title) is dict:
                item_dict = dict()
                # item_dict['日期'] = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%m月%d日')
                item_dict['日期'] = self.date_time
                item_dict['品牌'] = title['brand']
                item_dict['交易指数'] = title['tradeIndex']
                item_dict['支付商品数'] = title['payItemCnt']
                item_dict['客单价'] = title['payPct']
                item_dict['支付转化率'] = title['payRate']
                item_dict['访客数'] = title['uv']
                item_dict['搜索点击人数'] = title['searchUvCnt']
                item_dict['收藏人数'] = title['favBuyerCnt']
                item_dict['加购人数'] = title['addCartUserCnt']
                item_dict['卖家数'] = title['sellerCnt']
                item_dict['被支付卖家数'] = title['paySellerCnt']
                item_dict['重点卖家数'] = title['majorSellerCnt']
                item_dict['重点商品数'] = title['majorItemCnt']
                item_dict['每日累计区分'] = title['index']
                if 'washer' == title['cate']:
                    self.row_washer += 1
                    item_list = list(item_dict.keys())
                    for item_no in range(1, len(item_list) + 1):
                        self.ws_washer_add.cell(row=1, column=item_no, value=item_list[item_no - 1])
                        self.ws_washer_add.cell(row=self.row_washer, column=item_no, value=item_dict[item_list[item_no - 1]])
                    data = dict()
                    data['标识'] = title['index']
                    # data['日期'] = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%m月%d日')
                    data['日期'] = self.date_time
                    data['品牌'] = title['brand']
                    data['客单价'] = title['payPct']
                    data['支付转化率'] = str(round(title['payRate'] * 100, 2)) + '%'
                    data['访客数'] = title['uv']
                    data['销售额'] = title['payPct'] * title['payRate'] * title['uv']
                    print(data)
                    data_keys = list(data.keys())
                    print(data_keys)
                    if '1' == data['标识']:
                        if 'Haier/海尔' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 2, column=i, value=data[data_keys[i]])
                        elif 'Littleswan/小天鹅' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 5, column=i, value=data[data_keys[i]])
                        elif 'Midea/美的' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 6, column=i, value=data[data_keys[i]])
                        elif 'Sanyo/三洋' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 7, column=i, value=data[data_keys[i]])
                        elif 'TCL' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 8, column=i, value=data[data_keys[i]])
                        elif 'Panasonic/松下' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 9, column=i, value=data[data_keys[i]])
                        elif 'SIEMENS/西门子' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 10, column=i, value=data[data_keys[i]])
                        elif 'Leader/统帅' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 11, column=i, value=data[data_keys[i]])
                        elif 'Bosch/博世' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 12, column=i, value=data[data_keys[i]])
                        elif 'Royalstar/荣事达' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 13, column=i, value=data[data_keys[i]])
                        elif 'Whirlpool/惠而浦' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_washer.cell(row=self.maxRow_washer + 14, column=i, value=data[data_keys[i]])

                    elif '4' == data['标识']:
                        if 'Haier/海尔' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 2, column=i, value=data[data_keys[i - 4]])
                        elif 'Littleswan/小天鹅' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 5, column=i, value=data[data_keys[i - 4]])
                        elif 'Midea/美的' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 6, column=i, value=data[data_keys[i - 4]])
                        elif 'Sanyo/三洋' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 7, column=i, value=data[data_keys[i - 4]])
                        elif 'TCL' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 8, column=i, value=data[data_keys[i - 4]])
                        elif 'Panasonic/松下' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 9, column=i, value=data[data_keys[i - 4]])
                        elif 'SIEMENS/西门子' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 10, column=i, value=data[data_keys[i - 4]])
                        elif 'Leader/统帅' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 11, column=i, value=data[data_keys[i - 4]])
                        elif 'Bosch/博世' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 12, column=i, value=data[data_keys[i - 4]])
                        elif 'Royalstar/荣事达' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 13, column=i, value=data[data_keys[i - 4]])
                        elif 'Whirlpool/惠而浦' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_washer.cell(row=self.maxRow_washer + 14, column=i, value=data[data_keys[i - 4]])

                    self.ws_washer.cell(row=self.maxRow_washer + 3, column=1, value=data['日期'])
                    self.ws_washer.cell(row=self.maxRow_washer + 3, column=2, value='美的系')

                    self.ws_washer.cell(row=self.maxRow_washer + 4, column=1, value=data['日期'])
                    self.ws_washer.cell(row=self.maxRow_washer + 4, column=2, value='惠而浦系')

                    pass
                elif 'fridge' == title['cate']:
                    self.row_fridge += 1
                    item_list = list(item_dict.keys())
                    for item_no in range(1, len(item_list) + 1):
                        self.ws_fridge_add.cell(row=1, column=item_no, value=item_list[item_no - 1])
                        self.ws_fridge_add.cell(row=self.row_fridge, column=item_no, value=item_dict[item_list[item_no - 1]])
                    data = dict()
                    data['标识'] = title['index']
                    # data['日期'] = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%m月%d日')
                    data['日期'] = self.date_time
                    data['品牌'] = title['brand']
                    data['客单价'] = title['payPct']
                    data['支付转化率'] = str(round(title['payRate'] * 100, 2)) + '%'
                    data['访客数'] = title['uv']
                    data['销售额'] = title['payPct'] * title['payRate'] * title['uv']
                    print(data)
                    data_keys = list(data.keys())
                    print(data_keys)

                    if '标识' in data and '1' == data['标识']:
                        if 'Haier/海尔' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 2, column=i, value=data[data_keys[i]])
                        elif 'SIEMENS/西门子' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 3, column=i, value=data[data_keys[i]])
                        elif 'Midea/美的' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 4, column=i, value=data[data_keys[i]])
                        elif 'Samsung/三星' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 5, column=i, value=data[data_keys[i]])
                        elif 'Bosch/博世' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 6, column=i, value=data[data_keys[i]])
                        elif 'Whirlpool/惠而浦' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 7, column=i, value=data[data_keys[i]])
                        elif 'Royalstar/荣事达' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 8, column=i, value=data[data_keys[i]])
                        elif 'DIQUA/帝度' == data['品牌']:
                            for i in range(1, 7):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 9, column=i, value=data[data_keys[i]])

                    elif '标识' in data and '4' == data['标识']:
                        if 'Haier/海尔' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 2, column=i, value=data[data_keys[i - 4]])
                        elif 'SIEMENS/西门子' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 3, column=i, value=data[data_keys[i - 4]])
                        elif 'Midea/美的' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 4, column=i, value=data[data_keys[i - 4]])
                        elif 'Samsung/三星' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 5, column=i, value=data[data_keys[i - 4]])
                        elif 'Bosch/博世' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 6, column=i, value=data[data_keys[i - 4]])
                        elif 'Whirlpool/惠而浦' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 7, column=i, value=data[data_keys[i - 4]])
                        elif 'Royalstar/荣事达' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 8, column=i, value=data[data_keys[i - 4]])
                        elif 'DIQUA/帝度' == data['品牌']:
                            for i in range(7, 11):
                                self.ws_fridge.cell(row=self.maxRow_fridge + 9, column=i, value=data[data_keys[i - 4]])

        self.wb_washer.save(self.filename3)
        self.wb_fridge.save(self.filename4)
        # self.wb_washer.save('/Users/tiger007/Desktop/666/ratio/sanyo/02-5月洗衣机行业占比累计数据_%s.xlsx' % datetime.datetime.now().strftime('%Y%m%d'))
        # self.wb_fridge.save('/Users/tiger007/Desktop/666/ratio/sanyo/03-5月冰箱行业占比累计数据_%s.xlsx' % datetime.datetime.now().strftime('%Y%m%d'))

        # os.remove(self.filename1)
        # os.remove(self.filename2)
        return True