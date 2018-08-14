# -*- coding: utf-8 -*-
import os
import logging
import openpyxl
import datetime

from openpyxl.styles import Alignment


def singleton(cls, *args, **kw):
    instances = {}

    def _singleton(*args, **kw):
        if cls not in instances:
            instances[cls] = cls(*args, **kw)
        return instances[cls]

    return _singleton


@singleton
class SaveGmv(object):
    def __init__(self, download_dir):

        self.filename = download_dir + '销售明细_日_GMV口径_%s.xlsx' % (datetime.date.today() - datetime.timedelta(days=1)).strftime('%Y%m')
        if os.path.exists(self.filename):
            logging.info('cunzai ' * 30)
            self.wb = openpyxl.load_workbook(self.filename)

        else:
            logging.info('new ' * 30)
            self.wb = openpyxl.Workbook()

        try:
            logging.info('-' * 30)
            self.ws = self.wb[(datetime.date.today() - datetime.timedelta(days=1)).strftime('%Y%m%d')]
        except:
            logging.info('+' * 30)
            self.ws = self.wb.active
            logging.info(self.ws.title)
            if 'Sheet' == self.ws.title:
                self.ws.title = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%Y%m%d')
            else:
                # print(self.ws.title)
                self.ws = self.wb.create_sheet(title=(datetime.date.today() - datetime.timedelta(days=1)).strftime('%Y%m%d'), index=0)
                # self.ws = self.wb.active
                # if
                # self.ws.title = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%Y%m%d')
                self.ws = self.wb[(datetime.date.today() - datetime.timedelta(days=1)).strftime('%Y%m%d')]

            a1 = self.ws.cell(row=1, column=1, value='品牌')
            a1.alignment = Alignment(horizontal='center', vertical='center')
            self.ws.cell(row=1, column=2, value='型号').alignment = Alignment(horizontal='center', vertical='center')
            self.ws.cell(row=1, column=3, value='品类').alignment = Alignment(horizontal='center', vertical='center')
            self.ws.cell(row=1, column=4, value=(datetime.date.today() - datetime.timedelta(days=1)).strftime('%m月%d日')).alignment = Alignment(horizontal='center', vertical='center')
            self.ws.cell(row=2, column=4, value='数量').alignment = Alignment(horizontal='center', vertical='center')
            self.ws.cell(row=2, column=5, value='金额').alignment = Alignment(horizontal='center', vertical='center')
            self.ws.cell(row=1, column=6, value='店铺').alignment = Alignment(horizontal='center', vertical='center')

            # 品牌单元格合并
            self.ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
            # 型号单元格合并
            self.ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
            # 品类单元格合并
            self.ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
            # 日期单元格合并
            self.ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
            # 店铺点元格合并
            self.ws.merge_cells(start_row=1, start_column=6, end_row=2, end_column=6)





    def saveGmv(self, data):
        if data:
            logging.info(str(data))
            maxRow = self.ws.max_row
            maxCol = self.ws.max_column

            key_list = list(data.keys())

            try:
                shop_name = data['shop_name']
                key_list.remove('shop_name')
            except Exception as err:
                logging.error(err)
            else:
                for key in key_list:
                    maxRow += 1
                    res = data[key]
                    if 1 == len(res):
                        self.ws.cell(row=maxRow, column=1, value=shop_name)
                        self.ws.cell(row=maxRow, column=2, value=key)
                        self.ws.cell(row=maxRow, column=3, value=res[0]['type'])
                        self.ws.cell(row=maxRow, column=4, value=res[0]['buyNum'])
                        self.ws.cell(row=maxRow, column=5, value=res[0]['totalPrice'])
                        self.ws.cell(row=maxRow, column=6, value=res[0]['nick'])

                    elif len(res) > 1:
                        for num_price in res:
                            self.ws.cell(row=maxRow, column=1, value=shop_name)
                            self.ws.cell(row=maxRow, column=2, value=key)
                            self.ws.cell(row=maxRow, column=3, value=num_price['type'])
                            self.ws.cell(row=maxRow, column=4, value=num_price['buyNum'])
                            self.ws.cell(row=maxRow, column=5, value=num_price['totalPrice'])
                            self.ws.cell(row=maxRow, column=6, value=num_price['nick'])
                            maxRow += 1
                        maxRow -= 1


            self.wb.save(self.filename)

